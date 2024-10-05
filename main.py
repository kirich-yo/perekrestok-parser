import requests
from bs4 import BeautifulSoup
import itertools
import re
import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from pyfiglet import Figlet
from rich.console import Console
from rich.progress import Progress


CATALOG_SELECTOR = '#app > div > main > div > div > div > div.sc-dQoVA.bFTQaI.spinner-container-wrapper > div'
SUBCATEGORY_SELECTOR = 'div > div > div > div > div > div.catalog-content-group__list > div > div'
SUBCATEGORY_ITEM_SELECTOR = 'div.product-list-line > div.product-card-wrapper'
ITEM_NAME_SELECTOR = 'div.product-card__content > div.product-card__title-wrapper > div.product-card__title'
ITEM_PRICE_SELECTOR = 'div.product-card__content > div.product-card__control'
ITEM_IMG_SELECTOR = 'div.product-card__image-section > div.product-card__image-wrapper > img'
PRICE_REGEXP_PATTERN = r'(\d+),(\d{2})'


def parse_price(price_string):
    price = re.search(PRICE_REGEXP_PATTERN, price_string)

    if not price:
        return 0.0

    price = '.'.join(price.groups())
    return float(price)


def as_text(value):
    if value is None:
        return ""
    return str(value)


def save_to_xlsx(items):
    workbook = Workbook()
    sheet = workbook.active

    header_style = NamedStyle(name="header")
    header_style.font = Font(bold=True)
    header_style.border = Border(bottom=Side(border_style="thin"))
    header_style.alignment = Alignment(horizontal="center", vertical="center")

    sheet['A1'] = 'ИД'
    sheet['B1'] = 'Наименование'
    sheet['C1'] = 'Цена'
    for cell in sheet[1]:
        cell.style = header_style

    for i in range(len(items)):
        sheet[f'A{i+2}'] = i
        sheet[f'B{i+2}'] = items[i]['name']
        sheet[f'C{i+2}'] = items[i]['price']
        sheet[f'C{i+2}'].number_format = '#,##0.00 ₽'

    for column_cells in sheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 5

    workbook.save(filename='Goods.xlsx')


def mkdir_if_not_exists(a_dir):
    if not os.path.isdir(a_dir):
        os.mkdir(a_dir)


def download_image(item, id, dst_dir):
    response = requests.get(item['img'])
    with open(os.path.join(dst_dir, f'{id}.jpg'), 'wb') as f:
        f.write(response.content)
        


def parse_perekrestok(category):
    response = requests.get(f'https://www.perekrestok.ru/{category}')
    soup = BeautifulSoup(response.text, 'lxml')

    catalog = soup.select_one(CATALOG_SELECTOR)
    subcategories = [i for i in catalog.children][:-1] # The last tag does not contain useful information and is not being used in the script.
    subcategory_items = [list(itertools.chain.from_iterable(i.select(SUBCATEGORY_ITEM_SELECTOR) for i in subcategory.select(SUBCATEGORY_SELECTOR))) for subcategory in subcategories]
    all_items = list(itertools.chain.from_iterable(subcategory_items))
    all_items = [
        {
            'name': item.select_one(ITEM_NAME_SELECTOR).text,
            'price': parse_price(item.select_one(ITEM_PRICE_SELECTOR).text),
            'img': item.select_one(ITEM_IMG_SELECTOR).get('src')
        }
        for item in all_items
    ]

    return all_items


def main():
    if len(sys.argv) < 2:
        print('Perekrestok Parser by kirich_yo')
        print(f'Usage: {sys.argv[0]} [category]')
        print(f'Example: {sys.argv[0]} /cat/c/658/deserty-i-sneki')
        exit(1)

    figlet = Figlet(font='slant')
    print(figlet.renderText('PEREKRESTOK PARSER'))
    console = Console()

    try:
        with console.status('Fetching data from the server: [cyan link {0}]{0}[/]'.format(f'https://www.perekrestok.ru{sys.argv[1]}')):
            parsed_items = parse_perekrestok(sys.argv[1])
            if (not len(parsed_items)):
                console.print('[red bold]Failed to get all items. Try again later')
                exit(1)
        with console.status('Saving data to the Excel spreadsheet: [cyan]Goods.xlsx[/]'):
            save_to_xlsx(parsed_items)

        mkdir_if_not_exists('Pictures')

        with Progress() as progress:
            count = 0
            download_task = progress.add_task("Downloading all images...", total=len(parsed_items))
            while not progress.finished:
                download_image(parsed_items[count], count, 'Pictures')
                progress.update(download_task, advance=1)
                count += 1
    except requests.exceptions.ConnectionError as e:
        console.print(f'⛔ [red bold]An error occured while fetching data[/]: {str(e)}')
        console.print('Check your connection.')
        exit(1)
    except KeyboardInterrupt:
        console.print(f'✅ [green bold]Process aborted.')
        exit(1)
    except Exception as e:
        console.print_exception()
        exit(1)

    console.print('✅ [green bold]All tasks done')


if __name__ == '__main__':
    main()